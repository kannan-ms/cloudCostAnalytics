"""
File Parser Service - Parse CSV and Excel files for cost data
Handles file validation, parsing, and data extraction
REWRITTEN: Removed pandas dependency for better compatibility
"""

import io
import csv
import json
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

# Optional import for Excel
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# Supported file extensions
ALLOWED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
_LAST_PARSE_SUMMARY: Dict[str, Any] = {"dropped_rows": 0, "sample_errors": []}


def get_last_parse_summary() -> Dict[str, Any]:
    """Return parse diagnostics from the most recent parse operation."""
    return dict(_LAST_PARSE_SUMMARY)


def is_allowed_file(filename: str) -> bool:
    """Check if file extension is allowed."""
    return any(filename.lower().endswith(ext) for ext in ALLOWED_EXTENSIONS)


def validate_file_size(file_size: int) -> Tuple[bool, Optional[str]]:
    """Validate file size."""
    if file_size > MAX_FILE_SIZE:
        return False, f"File size exceeds maximum limit of {MAX_FILE_SIZE / (1024*1024)}MB"
    return True, None


def parse_date(date_value: Any) -> Optional[datetime]:
    """
    Parse various date formats to datetime object.
    Returns None if parsing fails or value is empty/None.
    """
    if date_value is None or date_value == '':
        return None
    
    # If already datetime
    if isinstance(date_value, datetime):
        return date_value
    
    # If integer, could be Excel serial date or Unix timestamp
    if isinstance(date_value, (int, float)):
        # Try as Unix timestamp (seconds since epoch)
        if 0 < date_value < 10000000000:  # Reasonable timestamp range
            try:
                return datetime.utcfromtimestamp(date_value)
            except (ValueError, OSError, OverflowError):
                pass
        # Try as Excel serial date (1900-01-01 is 1)
        if 1 < date_value < 100000:
            try:
                return datetime(1900, 1, 1) + __import__('datetime').timedelta(days=date_value - 2)
            except (ValueError, OverflowError):
                pass
    
    # If string, try various formats
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str:
            return None
            
        date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M:%S.%fZ',
            '%m-%d-%Y',
            '%d-%m-%Y',
            '%d-%m-%Y %H:%M',
            '%m-%d-%Y %H:%M',
            '%Y%m%d',  # YYYYMMDD format
            '%d-%b-%Y',  # 01-Jan-2024
            '%d %b %Y',  # 01 Jan 2024
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
    
    # Failed to parse
    return None


def normalize_column_name(col: str) -> str:
    """Normalize column names for flexible matching."""
    if not col:
        return ""
    return str(col).lower().strip().replace(' ', '_').replace('-', '_')


def map_columns(headers: List[str]) -> Dict[str, str]:
    """
    Map file headers to expected field names.
    """
    column_mapping = {}
    normalized_cols = {normalize_column_name(col): col for col in headers}
    
    # Define possible column name variations
    field_mappings = {
        'provider': ['provider', 'cloud_provider', 'cloud', 'vendor'],
        'cloud_account_id': ['cloud_account_id', 'account_id', 'account', 'account_number', 'subscriptionname', 'subscription_name', 'subscription_id', 'project_id'],
        'service_name': ['service_name', 'service', 'product', 'product_name', 'resource_type', 'service_description', 'sku_description', 'metercategory'],
        'consumed_service': ['consumedservice', 'consumed_service'],
        'resource_id': ['resource_id', 'resource', 'instance_id', 'instance', 'resourcename', 'resource_name'],
        'resource_group': ['resourcegroup', 'resource_group'],
        'region': ['region', 'location', 'availability_zone', 'zone', 'region/zone', 'resourcelocation', 'resource_location'],
        'usage_quantity': ['usage_quantity', 'quantity', 'usage', 'amount'],
        'usage_unit': ['usage_unit', 'unit', 'measurement_unit'],
        'cost': ['cost', 'charge', 'price', 'total_cost', 'billed_cost', 'unrounded_cost_($)', 'rounded_cost_($)', 'total_cost_(inr)', 'costinbillingcurrency', 'pretaxcost'],
        'currency': ['currency', 'currency_code', 'billingcurrency', 'billing_currency'],
        'usage_start_date': ['usage_start_date', 'start_date', 'from_date', 'begin_date', 'period_start', 'usage_start_time', 'start_time', 'date', 'usage_date', 'usagedatetime', 'usagestartdate'],
        'usage_end_date': ['usage_end_date', 'end_date', 'to_date', 'finish_date', 'period_end', 'usage_end_time', 'end_time'],
        'tags': ['tags', 'labels', 'metadata'],
    }
    
    # Try to match columns
    for field, possible_names in field_mappings.items():
        for possible_name in possible_names:
            normalized_possible = normalize_column_name(possible_name)
            if normalized_possible in normalized_cols:
                column_mapping[field] = normalized_cols[normalized_possible]
                break
    
    return column_mapping


def parse_tags(tag_value: Any) -> Dict:
    """Parse tags from various formats."""
    if not tag_value:
        return {}
    
    if isinstance(tag_value, dict):
        return tag_value
    
    if isinstance(tag_value, str):
        # Try JSON format
        try:
            return json.loads(tag_value)
        except:
            # Try key:value,key:value format
            tags = {}
            try:
                pairs = tag_value.split(',')
                for pair in pairs:
                    if ':' in pair:
                        key, val = pair.split(':', 1)
                        tags[key.strip()] = val.strip()
            except:
                pass
            return tags
    
    return {}


def _infer_provider(consumed_service: Optional[str] = None, service_name: Optional[str] = None) -> str:
    """
    Infer cloud provider from consumed_service or service_name.
    Azure services start with 'Microsoft.', AWS with 'Amazon'/'AWS', GCP with 'Google'.
    If no clear match, returns 'Other' instead of defaulting incorrectly.
    """
    for hint in [consumed_service, service_name]:
        if not hint:
            continue
        h = str(hint).lower().strip()
        if h.startswith('microsoft.') or 'azure' in h:
            return 'Azure'
        if h.startswith('amazon') or h.startswith('aws') or 'aws' in h:
            return 'AWS'
        if h.startswith('google') or 'gcp' in h or h.startswith('cloud.'):
            return 'GCP'
    return 'Other'  # Unknown provider - don't guess


def extract_cost_records(rows: List[Dict[str, Any]], column_mapping: Dict[str, str], parse_errors: Optional[List[Dict[str, Any]]] = None) -> List[Dict]:
    """
    Extract cost records from list of dictionaries (rows) using column mapping.
    """
    records = []
    
    for row_index, row in enumerate(rows, start=1):
        try:
            record = {}
            
            # Helper to safely get value from row
            def get_val(field_key):
                col_name = column_mapping.get(field_key)
                if col_name and col_name in row:
                    return row[col_name]
                return None
            
            # Provider — explicit column or inferred from consumed_service
            provider = get_val('provider')
            consumed_service = get_val('consumed_service')
            if provider:
                record['provider'] = str(provider).strip().lower()
            else:
                inferred = _infer_provider(
                    consumed_service=str(consumed_service) if consumed_service else None,
                    service_name=str(get_val('service_name')) if get_val('service_name') else None
                )
                record['provider'] = inferred.lower()
            
            # Service name (required) - must not be empty
            service = get_val('service_name')
            if not service or not str(service).strip():
                if parse_errors is not None:
                    parse_errors.append({'row': row_index, 'field': 'service_name', 'error': 'Missing or empty service name'})
                continue
            record['service_name'] = str(service).strip()
            
            # Cost (required) - must be numeric and reasonable
            cost = get_val('cost')
            if cost is None or cost == '':
                if parse_errors is not None:
                    parse_errors.append({'row': row_index, 'field': 'cost', 'error': 'Missing cost value'})
                continue
            
            try:
                # Handle currency symbols and commas if valid string
                if isinstance(cost, str):
                    cost = cost.replace('$', '').replace('\u20ac', '').replace('\u00a3', '').replace(',', '').strip()
                cost_float = float(cost)
                # Allow negative costs (refunds) and zero costs (free tier), but catch NaN/Inf
                if not (cost_float == cost_float):  # NaN check
                    raise ValueError("Invalid cost: NaN")
                record['cost'] = cost_float
            except (ValueError, TypeError) as e:
                if parse_errors is not None:
                    parse_errors.append({'row': row_index, 'field': 'cost', 'error': f'Invalid numeric value: {str(e)}'})
                continue
            
            # Dates (required)
            start_date = parse_date(get_val('usage_start_date'))
            if not start_date:
                if parse_errors is not None:
                    parse_errors.append({'row': row_index, 'field': 'usage_start_date', 'error': 'Missing or invalid start date'})
                continue
            record['usage_start_date'] = start_date.isoformat()
            
            end_date = parse_date(get_val('usage_end_date'))
            if end_date:
                record['usage_end_date'] = end_date.isoformat()
            else:
                # If no end date column, default to start date
                record['usage_end_date'] = start_date.isoformat()
            
            # Optional fields
            account_id = get_val('cloud_account_id')
            if account_id and str(account_id).strip():
                record['cloud_account_id'] = str(account_id).strip()
            
            resource_id = get_val('resource_id')
            if resource_id and str(resource_id).strip():
                record['resource_id'] = str(resource_id).strip()
            
            resource_group = get_val('resource_group')
            if resource_group and str(resource_group).strip():
                record['tags'] = record.get('tags', {})
                record['tags']['resource_group'] = str(resource_group).strip()
            
            region = get_val('region')
            if region and str(region).strip():
                record['region'] = str(region).strip()
            
            quantity = get_val('usage_quantity')
            if quantity is not None:
                try:
                    qty_float = float(quantity)
                    if qty_float == qty_float:  # Check for NaN
                        record['usage_quantity'] = qty_float
                except (ValueError, TypeError):
                    pass
            
            unit = get_val('usage_unit')
            if unit and str(unit).strip():
                record['usage_unit'] = str(unit).strip()
            
            currency = get_val('currency')
            if currency and str(currency).strip():
                record['currency'] = str(currency).strip().upper()
            
            tags = parse_tags(get_val('tags'))
            if tags:
                record['tags'] = {**record.get('tags', {}), **tags}
            
            # Record should have all required fields
            if all(k in record for k in ['provider', 'service_name', 'cost', 'usage_start_date', 'usage_end_date']):
                records.append(record)
            else:
                if parse_errors is not None:
                    missing = [k for k in ['provider', 'service_name', 'cost', 'usage_start_date', 'usage_end_date'] if k not in record]
                    parse_errors.append({'row': row_index, 'field': 'record', 'error': f'Missing required fields: {missing}'})
        
        except Exception as e:
            # Skip problematic rows and log the error
            if parse_errors is not None:
                parse_errors.append({'row': row_index, 'field': 'row', 'error': f'Row parsing failed: {str(e)}'})
            continue
    
    return records


def parse_csv_file(file_content: bytes) -> Tuple[bool, Any]:
    """
    Parse CSV file and extract cost records.
    """
    try:
        global _LAST_PARSE_SUMMARY
        _LAST_PARSE_SUMMARY = {"dropped_rows": 0, "sample_errors": []}

        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        decoded_content = None
        
        for encoding in encodings:
            try:
                decoded_content = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if decoded_content is None:
            return False, "Unable to decode CSV file. Please ensure it's a valid CSV with UTF-8 encoding."
        
        # Parse CSV to list of dicts
        f = io.StringIO(decoded_content)
        reader = csv.DictReader(f)
        
        # Get headers from the reader
        if not reader.fieldnames:
             return False, "CSV file is empty or has no headers"
             
        rows = list(reader)
        if not rows:
             return False, "CSV file contains no data rows"

        # Map columns
        column_mapping = map_columns(reader.fieldnames)
        
        # Check for minimum required columns (provider & end_date can be inferred)
        required_fields = ['service_name', 'cost', 'usage_start_date']
        missing_fields = [f for f in required_fields if f not in column_mapping]
        
        if missing_fields:
            available = ', '.join(column_mapping.keys())
            return False, f"Missing required columns: {', '.join(missing_fields)}. Detected columns: {available}. Please ensure your CSV has at minimum: service_name (or MeterCategory), cost (or CostInBillingCurrency), and a date column."
        
        # Extract records
        parse_errors = []
        records = extract_cost_records(rows, column_mapping, parse_errors=parse_errors)
        
        if not records:
            return False, "No valid records could be extracted from the CSV file. Please check that your data rows contain valid service names, costs, and dates."

        _LAST_PARSE_SUMMARY = {
            "dropped_rows": len(parse_errors),
            "sample_errors": parse_errors[:5]
        }
        
        return True, records
        
    except Exception as e:
        return False, f"Error parsing CSV file: {str(e)}"


def parse_excel_file(file_content: bytes) -> Tuple[bool, Any]:
    """
    Parse Excel file and extract cost records.
    """
    if not openpyxl:
        return False, "openpyxl library is missing (required for .xlsx)"

    try:
        global _LAST_PARSE_SUMMARY
        _LAST_PARSE_SUMMARY = {"dropped_rows": 0, "sample_errors": []}

        # Load workbook
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        if sheet.max_row < 2:
            return False, "Excel file is empty"
            
        # Get headers (first row)
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else "")
            
        # Parse rows to list of dicts
        rows = []
        for row in sheet.iter_rows(min_row=2):
            row_dict = {}
            has_data = False
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    header = headers[idx]
                    row_dict[header] = cell.value
                    if cell.value is not None:
                        has_data = True
            if has_data:
                rows.append(row_dict)
        
        if not rows:
            return False, "Excel file contains no data rows"

        # Map columns
        column_mapping = map_columns(headers)
        
        # Check for minimum required columns (provider & end_date can be inferred)
        required_fields = ['service_name', 'cost', 'usage_start_date']
        missing_fields = [f for f in required_fields if f not in column_mapping]
        
        if missing_fields:
            available = ', '.join(column_mapping.keys())
            return False, f"Missing required columns: {', '.join(missing_fields)}. Detected columns: {available}. Please ensure your Excel has at minimum: service_name, cost, and a date column."
        
        # Extract records
        parse_errors = []
        records = extract_cost_records(rows, column_mapping, parse_errors=parse_errors)
        
        if not records:
            return False, "No valid records could be extracted from the Excel file. Please check that your data rows contain valid service names, costs, and dates."

        _LAST_PARSE_SUMMARY = {
            "dropped_rows": len(parse_errors),
            "sample_errors": parse_errors[:5]
        }
        
        return True, records
        
    except Exception as e:
        return False, f"Error parsing Excel file: {str(e)}"


def parse_file(filename: str, file_content: bytes) -> Tuple[bool, Any]:
    """
    Parse uploaded file (CSV or Excel) and extract cost records.
    """
    # Validate file extension
    if not is_allowed_file(filename):
        return False, f"File type not allowed. Supported formats: {', '.join(ALLOWED_EXTENSIONS)}"
    
    # Validate file size
    is_valid_size, error = validate_file_size(len(file_content))
    if not is_valid_size:
        return False, error
    
    # Parse based on file type
    if filename.lower().endswith('.csv'):
        return parse_csv_file(file_content)
    elif filename.lower().endswith(('.xlsx', '.xls')):
        # Note: .xls support is limited without xlrd, but modern systems use xlsx
        return parse_excel_file(file_content)
    else:
        return False, "Unsupported file format"
